"""M6 Excel / JSON 导出测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.db import init_db, save_batch
from app.exporter import TaskNotFoundError, export_excel, export_json
from app.exporter import _estimated_row_height, _safe_excel_text
from app.models import (
    BatchFileResult,
    BatchResult,
    DocumentType,
    ExtractionFailure,
    ExtractionStage,
    PaperRecord,
)


def sample_batch() -> BatchResult:
    success = BatchFileResult(
        path=Path("中文论文.pdf"),
        filename="中文论文.pdf",
        success=True,
        record=PaperRecord(
            title="可解释的论文信息抽取",
            authors=["张三", "Alice"],
            year=2026,
            venue="DemoConf",
            doc_type=DocumentType.CONFERENCE_PAPER,
            problem="如何稳定抽取论文结构化字段",
            method_name="PaperExtractor",
            experimental_conditions=["16-Gbaud 16QAM", "20 km SSMF"],
            main_results=["F1 提升 5%", "失败率下降 20%"],
            limitations="仅在电子版 PDF 上验证",
            summary="通过分层容错和批量调度生成论文对比表。",
        ),
        retry_count=1,
        total_tokens=321,
        duration_ms=456,
    )
    failure = BatchFileResult(
        path=Path("损坏论文.pdf"),
        filename="损坏论文.pdf",
        success=False,
        failure=ExtractionFailure(
            stage=ExtractionStage.PARSE,
            error_type="CorruptedDocumentError",
            error_msg="文件损坏",
        ),
        duration_ms=8,
    )
    return BatchResult(
        total_files=2,
        success_count=1,
        fail_count=1,
        total_tokens=321,
        duration_ms=500,
        files=[success, failure],
    )


def test_export_json_keeps_complete_structure_and_unicode(tmp_path: Path) -> None:
    db_path = tmp_path / "app.db"
    task_id = save_batch(sample_batch(), db_path)

    output = export_json(task_id, tmp_path / "nested" / "result.json", db_path=db_path)
    payload = json.loads(output.read_text(encoding="utf-8"))

    assert payload["task"]["id"] == task_id
    assert payload["task"]["total_files"] == 2
    assert payload["results"][0]["authors"] == ["张三", "Alice"]
    assert payload["results"][0]["experimental_conditions"] == [
        "16-Gbaud 16QAM",
        "20 km SSMF",
    ]
    assert payload["results"][0]["main_results"] == ["F1 提升 5%", "失败率下降 20%"]
    assert payload["failures"][0]["error_type"] == "CorruptedDocumentError"


def test_export_excel_has_result_failure_and_task_sheets(tmp_path: Path) -> None:
    db_path = tmp_path / "app.db"
    task_id = save_batch(sample_batch(), db_path)

    output = export_excel(task_id, tmp_path / "result.xlsx", db_path=db_path)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["论文结果", "失败记录", "任务概览"]
    result_sheet = workbook["论文结果"]
    failure_sheet = workbook["失败记录"]
    task_sheet = workbook["任务概览"]
    assert result_sheet.max_row == 2
    assert result_sheet["A2"].value == "中文论文.pdf"
    assert result_sheet["C2"].value == "张三\nAlice"
    assert result_sheet["I2"].value == "16-Gbaud 16QAM\n20 km SSMF"
    assert result_sheet["J2"].value == "F1 提升 5%\n失败率下降 20%"
    assert failure_sheet.max_row == 2
    assert failure_sheet["A2"].value == "损坏论文.pdf"
    assert failure_sheet["B2"].value == "parse"
    assert task_sheet.max_row == 2
    assert task_sheet["A2"].value == task_id
    assert task_sheet["C2"].value == 2
    assert task_sheet["D2"].value == 1
    assert task_sheet["E2"].value == 1


def test_excel_is_readable_and_has_navigation_formatting(tmp_path: Path) -> None:
    db_path = tmp_path / "app.db"
    task_id = save_batch(sample_batch(), db_path)

    output = export_excel(task_id, tmp_path / "result.xlsx", db_path=db_path)
    workbook = load_workbook(output)

    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].font.color.rgb == "00FFFFFF"
        assert sheet.sheet_view.showGridLines is False
        assert sheet.column_dimensions["A"].width >= 10


def test_excel_keeps_empty_failure_sheet_with_headers(tmp_path: Path) -> None:
    db_path = tmp_path / "app.db"
    batch = sample_batch()
    success_only = BatchResult(
        total_files=1,
        success_count=1,
        fail_count=0,
        total_tokens=batch.files[0].total_tokens,
        duration_ms=480,
        files=[batch.files[0]],
    )
    task_id = save_batch(success_only, db_path)

    output = export_excel(task_id, tmp_path / "result.xlsx", db_path=db_path)
    workbook = load_workbook(output)

    assert workbook["失败记录"].max_row == 1
    assert workbook["失败记录"]["A1"].value == "文件名"


@pytest.mark.parametrize("value", ["=1+1", "+SUM(A1:A2)", "-1+2", "@cmd"])
def test_excel_formula_like_text_is_escaped(value: str) -> None:
    assert _safe_excel_text(value) == f"'{value}"


def test_long_multiline_content_expands_row_height() -> None:
    height = _estimated_row_height(
        ["作者一\n作者二\n作者三\n作者四\n作者五"],
        [20],
    )

    assert height == 75


@pytest.mark.parametrize(
    ("function", "filename", "message"),
    [
        (export_json, "wrong.txt", ".json"),
        (export_excel, "wrong.csv", ".xlsx"),
    ],
)
def test_export_rejects_wrong_extension(
    tmp_path: Path,
    function: object,
    filename: str,
    message: str,
) -> None:
    db_path = tmp_path / "app.db"
    task_id = save_batch(sample_batch(), db_path)

    with pytest.raises(ValueError, match=message.replace(".", r"\.")):
        function(task_id, tmp_path / filename, db_path=db_path)  # type: ignore[operator]


@pytest.mark.parametrize("function,extension", [(export_json, "json"), (export_excel, "xlsx")])
def test_missing_task_has_clear_error(
    tmp_path: Path,
    function: object,
    extension: str,
) -> None:
    db_path = tmp_path / "app.db"
    init_db(db_path)

    with pytest.raises(TaskNotFoundError, match="task_id=999"):
        function(999, tmp_path / f"missing.{extension}", db_path=db_path)  # type: ignore[operator]
