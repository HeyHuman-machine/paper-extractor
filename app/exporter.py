"""M6 导出：把 M5 中指定任务的结果导出为 Excel 或 JSON。"""

from __future__ import annotations

import json
import math
import unicodedata
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.db import get_failures, get_results, get_task


RESULT_COLUMNS = (
    ("filename", "文件名"),
    ("title", "标题"),
    ("authors", "作者"),
    ("year", "年份"),
    ("venue", "期刊 / 会议"),
    ("doc_type", "文档类型"),
    ("problem", "研究问题"),
    ("method_name", "方法名称"),
    ("experimental_conditions", "实验条件"),
    ("main_results", "主要结果"),
    ("limitations", "局限性"),
    ("summary", "摘要总结"),
    ("retry_count", "抽取重试次数"),
    ("tokens", "Token 数"),
    ("latency_ms", "耗时（毫秒）"),
)

FAILURE_COLUMNS = (
    ("filename", "文件名"),
    ("stage", "失败阶段"),
    ("error_type", "错误类型"),
    ("error_msg", "错误信息"),
    ("retry_count", "抽取重试次数"),
    ("created_at", "记录时间"),
    ("raw_output", "模型原始输出"),
)

TASK_COLUMNS = (
    ("id", "任务 ID"),
    ("status", "任务状态"),
    ("total_files", "文件总数"),
    ("success_count", "成功数"),
    ("fail_count", "失败数"),
    ("total_tokens", "Token 总数"),
    ("duration_ms", "总耗时（毫秒）"),
    ("created_at", "创建时间（UTC）"),
    ("finished_at", "完成时间（UTC）"),
)

_HEADER_FILL = PatternFill("solid", fgColor="244B64")
_FAILURE_HEADER_FILL = PatternFill("solid", fgColor="76505A")
_TASK_HEADER_FILL = PatternFill("solid", fgColor="496A74")
_EVEN_ROW_FILL = PatternFill("solid", fgColor="F1F6F9")
_THIN_BOTTOM = Border(bottom=Side(style="thin", color="D5E0E7"))


class TaskNotFoundError(LookupError):
    """请求导出的 task_id 在数据库中不存在。"""


def export_json(
    task_id: int,
    path: Path | str,
    *,
    db_path: Path | str | None = None,
) -> Path:
    """导出任务总览、成功结果和失败诊断，返回生成文件路径。"""

    task, results, failures = _load_task_data(task_id, db_path)
    output_path = _prepare_output_path(path, ".json")
    payload = {
        "task": task,
        "results": results,
        "failures": failures,
    }
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output_path


def export_excel(
    task_id: int,
    path: Path | str,
    *,
    db_path: Path | str | None = None,
) -> Path:
    """导出可直接作为论文对比表使用的 Excel 工作簿。"""

    task, results, failures = _load_task_data(task_id, db_path)
    output_path = _prepare_output_path(path, ".xlsx")

    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "论文结果"
    failure_sheet = workbook.create_sheet("失败记录")
    task_sheet = workbook.create_sheet("任务概览")

    _write_table(result_sheet, RESULT_COLUMNS, results, _HEADER_FILL)
    _write_table(failure_sheet, FAILURE_COLUMNS, failures, _FAILURE_HEADER_FILL)
    _write_table(task_sheet, TASK_COLUMNS, [task], _TASK_HEADER_FILL)

    workbook.save(output_path)
    return output_path


def _load_task_data(
    task_id: int,
    db_path: Path | str | None,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    task = get_task(task_id, db_path)
    if task is None:
        raise TaskNotFoundError(f"找不到 task_id={task_id} 的任务")
    return task, get_results(task_id, db_path), get_failures(task_id, db_path)


def _prepare_output_path(path: Path | str, expected_suffix: str) -> Path:
    output_path = Path(path)
    if output_path.suffix.lower() != expected_suffix:
        raise ValueError(f"导出路径必须以 {expected_suffix} 结尾")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


def _write_table(
    sheet: Worksheet,
    columns: tuple[tuple[str, str], ...],
    rows: list[dict[str, Any]],
    header_fill: PatternFill,
) -> None:
    headers = [label for _field, label in columns]
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(row.get(field)) for field, _label in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(sheet.max_row, 1)}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 30
    _set_adaptive_widths(sheet, columns)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = _EVEN_ROW_FILL
        for cell in sheet[row_number]:
            cell.font = Font(name="Microsoft YaHei", size=10, color="263746")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BOTTOM
        widths = [
            sheet.column_dimensions[get_column_letter(index)].width or 10
            for index in range(1, len(columns) + 1)
        ]
        values = [
            sheet.cell(row=row_number, column=index).value
            for index in range(1, len(columns) + 1)
        ]
        sheet.row_dimensions[row_number].height = _estimated_row_height(values, widths)


def _set_adaptive_widths(
    sheet: Worksheet,
    columns: tuple[tuple[str, str], ...],
) -> None:
    width_caps = {
        "filename": 28,
        "title": 38,
        "authors": 30,
        "problem": 46,
        "experimental_conditions": 38,
        "main_results": 50,
        "limitations": 42,
        "summary": 50,
        "error_msg": 48,
        "raw_output": 50,
    }
    for column_index, (field, _label) in enumerate(columns, start=1):
        values = (sheet.cell(row=row, column=column_index).value for row in range(1, sheet.max_row + 1))
        longest = max((_display_width(value) for value in values), default=10)
        cap = width_caps.get(field, 24)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            cap, max(10, longest + 2)
        )


def _excel_value(value: Any) -> str | int | float | None:
    if isinstance(value, list):
        return _safe_excel_text("\n".join(str(item) for item in value))
    if isinstance(value, str):
        return _safe_excel_text(value)
    if value is None or isinstance(value, (int, float)):
        return value
    return _safe_excel_text(str(value))


def _safe_excel_text(value: str) -> str:
    """避免论文文本被 Excel 当成公式执行。"""

    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _display_width(value: Any) -> int:
    text = "" if value is None else str(value)
    lines = text.splitlines() or [""]
    return max(_line_display_width(line) for line in lines)


def _estimated_row_height(values: list[Any], widths: list[float]) -> float:
    """按显式换行和列宽估算行高，同时限制极端长文本占用的高度。"""

    estimated_lines = 1
    for value, width in zip(values, widths, strict=True):
        text = "" if value is None else str(value)
        usable_width = max(1, width - 2)
        line_count = sum(
            max(1, math.ceil(_line_display_width(line) / usable_width))
            for line in (text.splitlines() or [""])
        )
        estimated_lines = max(estimated_lines, line_count)
    return min(180, max(42, estimated_lines * 15))


def _line_display_width(text: Iterable[str]) -> int:
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in text)
